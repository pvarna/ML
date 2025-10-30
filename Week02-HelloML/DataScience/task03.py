import numpy as np
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl import load_workbook

TELECOM_PATH = "../../DATA/telecom_churn_clean.csv"
OUTPUT_XLSX = "data_audit.xlsx"


def build_summary(df_telecom):
    summary = df_telecom.describe().T

    num_missing = df_telecom.isna().sum()
    pct_missing = (num_missing / len(df_telecom))
    num_unique = df_telecom.nunique()
    pct_unique = (num_unique / len(df_telecom))

    summary["NUM_MISSING"] = num_missing
    summary["%_MISSING"] = pct_missing
    summary["NUM_UNIQUE"] = num_unique
    summary["%_UNIQUE"] = pct_unique
    summary["UNIVARIATE ANALYSIS COMMENTS"] = ""
    summary["MULTIVARIATE ANALYSIS COMMENTS"] = ""

    summary = summary.round(2)
    return summary


def calculate_value_counts(df_telecom):
    return {
        col: df_telecom[col].value_counts().sort_index()
        for col in df_telecom.columns
    }


def write_tables(summary, features):
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="overview")
        for feature_name, value_counts in features.items():
            value_counts.to_excel(writer, sheet_name=feature_name)


def plot_feature(feature_name, value_counts, df_telecom):
    plt.figure(figsize=(6, 4))

    if len(value_counts.index) < 10:
        df_plot = value_counts.reset_index()
        df_plot.columns = [feature_name, "count"]
        sns.barplot(data=df_plot, x=feature_name, y="count")
        plt.title(f"Bar chart of {feature_name}")
        plt.ylabel("Count")
        plt.xlabel(feature_name)
    else:
        sns.histplot(df_telecom[feature_name], bins=30)
        plt.title(f"Histogram of {feature_name}")
        plt.ylabel("Frequency")
        plt.xlabel(feature_name)

    plt.tight_layout()
    img_path = f"assets/{feature_name}.png"
    plt.savefig(img_path)
    plt.close()

    return img_path


def insert_img_into_sheet(workbook, feature_name, img_path):
    if feature_name in workbook.sheetnames:
        worksheet = workbook[feature_name]
        worksheet.add_image(Image(img_path), "E1")


def plot_overwiew_pairplot(df_telecom):
    pair = sns.pairplot(data=df_telecom, hue="churn", diag_kind="kde")
    img_path = f"assets/overview.png"
    pair.savefig(img_path)
    plt.close()
    return img_path


def insert_overview_img(workbook, img_path):
    workbook["overview"].add_image(Image(img_path), "A25")


def main():
    df_telecom = pd.read_csv(TELECOM_PATH, delimiter=',', index_col=0)

    summary = build_summary(df_telecom)
    features = calculate_value_counts(df_telecom)

    write_tables(summary, features)
    workbook = load_workbook(OUTPUT_XLSX)
    for feature_name, value_counts in features.items():
        feature_img_path = plot_feature(feature_name, value_counts, df_telecom)
        insert_img_into_sheet(workbook, feature_name, feature_img_path)

    overview_img_path = plot_overwiew_pairplot(df_telecom)
    insert_overview_img(workbook, overview_img_path)
    workbook.save(OUTPUT_XLSX)

    # Google Sheet with added comments - https://docs.google.com/spreadsheets/d/1eDuaXD4sAXf9v5BNwyfciUiZaB-xvB3pbgbXbaWuyP4/edit?usp=sharing


if __name__ == '__main__':
    main()
