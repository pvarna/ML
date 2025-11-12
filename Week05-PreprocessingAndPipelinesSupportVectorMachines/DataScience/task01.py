import os
import json
import numpy as np
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
import time

DATASET_PATH = os.path.join("..", "..", "DATA", "music_dirty.txt")
OUTPUT_XLSX = "data_audit.xlsx"
ASSETS_DIR = "assets"
TARGET = "popularity"
NON_NUMERIC_FEATURE = "genre"


def log(msg):
    print(f"[{time.strftime('%H:%M:%S')}] {msg}")


def build_summary(df):
    summary = pd.DataFrame(index=df.columns)

    numeric_desc = df.describe().T
    summary = summary.join(numeric_desc, how="left")

    num_missing = df.isna().sum()
    pct_missing = num_missing / len(df)
    num_unique = df.nunique(dropna=True)
    pct_unique = num_unique / len(df)

    summary["NUM_MISSING"] = num_missing
    summary["%_MISSING"] = pct_missing
    summary["NUM_UNIQUE"] = num_unique
    summary["%_UNIQUE"] = pct_unique

    summary["UNIVARIATE ANALYSIS COMMENTS"] = ""
    summary["MULTIVARIATE ANALYSIS COMMENTS"] = ""

    summary = summary.round(2)

    log("Summary table built.")

    return summary


def calculate_value_counts(df):
    log("Calculating value counts for all columns...")
    return {col: df[col].value_counts().sort_index() for col in df.columns}


def write_tables(summary, features):
    log("Writing tables to Excel...")
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="overview")
        for feature_name, value_counts in features.items():
            value_counts.to_excel(writer, sheet_name=feature_name)

    log("Finished writing tables.")


def plot_feature(feature_name, value_counts, df):
    log(f"Plotting feature '{feature_name}'...")
    plt.figure(figsize=(6, 4))

    if len(value_counts.index) < 10:
        df_plot = value_counts.reset_index()
        df_plot.columns = [feature_name, "count"]
        sns.barplot(data=df_plot, x=feature_name, y="count")
        plt.title(f"Bar chart of {feature_name}")
        plt.ylabel("Count")
        plt.xlabel(feature_name)
    else:
        sns.histplot(df[feature_name], bins=30)
        plt.title(f"Histogram of {feature_name}")
        plt.ylabel("Frequency")
        plt.xlabel(feature_name)

    plt.tight_layout()
    img_path = f"assets/{feature_name}.png"
    plt.savefig(img_path)
    plt.close()

    log(f"  Saved plot: {img_path}")
    return img_path


def insert_img_into_sheet(workbook, feature_name, img_path):
    if feature_name in workbook.sheetnames:
        log(f"Inserting image into sheet '{feature_name}'...")
        worksheet = workbook[feature_name]
        worksheet.add_image(Image(img_path), "E1")


def plot_overwiew_pairplot(df):
    log("Creating overview pairplot... this may take a while.")
    pair = sns.pairplot(data=df, diag_kind="kde", kind="reg")
    img_path = os.path.join(ASSETS_DIR, "overview.png")
    pair.savefig(img_path)
    plt.close()
    log(f"Overview pairplot saved to {img_path}")
    return img_path


def plot_correlation_heatmap(df):
    log("Creating correlation heatmap...")

    corr = df.corr(numeric_only=True)

    plt.figure(figsize=(8, 6))
    sns.heatmap(
        corr,
        annot=True,
        fmt=".2f",
        annot_kws={"size": 5},
        cmap="coolwarm",
        square=True,
        cbar_kws={"shrink": 0.8},
    )
    plt.title("Correlation Heatmap")
    plt.tight_layout()

    img_path = os.path.join(ASSETS_DIR, "correlation_heatmap.png")
    plt.savefig(img_path)
    plt.close()

    log(f"Correlation heatmap saved to {img_path}")
    return img_path


def insert_imgs(workbook, overview_img_path, correlation_img_path):
    log("Inserting overview and correlation images into 'overview' sheet...")
    workbook["overview"].add_image(Image(correlation_img_path), "A30")
    workbook["overview"].add_image(Image(overview_img_path), "A60")
    log("Images inserted.")


def create_dataset(path):
    log(f"Loading dataset from {path}...")
    with open(path, "r") as f:
        data = json.load(f)

    log("Dataset loaded.")
    return pd.DataFrame(data)


def create_dummy_variables(df, non_numeric_feature):
    log(f"Creating dummy variables for '{non_numeric_feature}'...")
    dummies = pd.get_dummies(df[non_numeric_feature],
                             drop_first=True,
                             dtype=int)
    dummies = pd.concat([df, dummies], axis=1)
    dummies = dummies.drop(columns=[non_numeric_feature])
    log("Dummy variables created.")

    return dummies


def main():
    start_time = time.time()
    log("===== Data Audit Script Started =====")

    df = create_dataset(DATASET_PATH)
    df = create_dummy_variables(df, NON_NUMERIC_FEATURE)

    log(f"Dataset shape: {df.shape}")

    os.makedirs(ASSETS_DIR, exist_ok=True)

    summary = build_summary(df)
    features = calculate_value_counts(df)

    write_tables(summary, features)
    workbook = load_workbook(OUTPUT_XLSX)
    for feature_name, value_counts in features.items():
        log(f"Plotting and inserting '{feature_name}'...")
        feature_img_path = plot_feature(feature_name, value_counts, df)
        insert_img_into_sheet(workbook, feature_name, feature_img_path)

    overview_img_path = plot_overwiew_pairplot(df)
    correlation_img_path = plot_correlation_heatmap(df)
    insert_imgs(workbook, overview_img_path, correlation_img_path)
    workbook.save(OUTPUT_XLSX)

    total_time = time.time() - start_time
    log(f"===== Script finished in {total_time:.2f} seconds =====")

    # data_audit - https://docs.google.com/spreadsheets/d/1HSV5wir92NMxP5HekZMNTkPHnTRoL1pdn-raWGXi2jc/edit?usp=sharing

if __name__ == '__main__':
    main()
