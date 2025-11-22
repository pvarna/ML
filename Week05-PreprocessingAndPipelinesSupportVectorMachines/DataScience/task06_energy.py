import os
import json
import numpy as np
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl import load_workbook

DATASET_PATH = os.path.join("..", "..", "DATA", "music_clean.csv")
OUTPUT_XLSX = "data_audit.xlsx"
ASSETS_DIR = "assets"
TARGET = "energy"


def build_summary(df):
    summary = pd.DataFrame(index=df.columns)

    numeric_desc = df.describe().T
    summary = summary.join(numeric_desc, how="left")

    num_missing = df.isna().sum()
    pct_missing = num_missing / len(df)
    num_unique = df.nunique(dropna=True)
    pct_unique = num_unique / len(df)

    summary["NUM_MISSING"] = num_missing
    summary["%_MISSING"] = pct_missing
    summary["NUM_UNIQUE"] = num_unique
    summary["%_UNIQUE"] = pct_unique

    summary["UNIVARIATE ANALYSIS COMMENTS"] = ""
    summary["MULTIVARIATE ANALYSIS COMMENTS"] = ""

    summary = summary.round(2)

    return summary


def calculate_value_counts(df):
    return {col: df[col].value_counts().sort_index() for col in df.columns}


def write_tables(summary, features):
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="overview")
        for feature_name, value_counts in features.items():
            value_counts.to_excel(writer, sheet_name=feature_name)


def plot_feature(feature_name, value_counts, df):
    plt.figure(figsize=(6, 4))

    if len(value_counts.index) < 10:
        df_plot = value_counts.reset_index()
        df_plot.columns = [feature_name, "count"]
        sns.barplot(data=df_plot, x=feature_name, y="count")
        plt.title(f"Bar chart of {feature_name}")
        plt.ylabel("Count")
        plt.xlabel(feature_name)
    else:
        sns.histplot(df[feature_name], bins=30)
        plt.title(f"Histogram of {feature_name}")
        plt.ylabel("Frequency")
        plt.xlabel(feature_name)

    plt.tight_layout()
    img_path = f"assets/{feature_name}.png"
    plt.savefig(img_path)
    plt.close()

    return img_path


def insert_img_into_sheet(workbook, feature_name, img_path):
    if feature_name in workbook.sheetnames:
        worksheet = workbook[feature_name]
        worksheet.add_image(Image(img_path), "E1")


def plot_overwiew_pairplot(df):
    pair = sns.pairplot(data=df, diag_kind="kde", kind="reg")
    img_path = os.path.join(ASSETS_DIR, "overview.png")
    pair.savefig(img_path)
    plt.close()
    return img_path


def plot_correlation_heatmap(df):
    corr = df.corr(numeric_only=True)

    plt.figure(figsize=(8, 6))
    sns.heatmap(
        corr,
        annot=True,
        fmt=".2f",
        annot_kws={"size": 5},
        cmap="coolwarm",
        square=True,
        cbar_kws={"shrink": 0.8},
    )
    plt.title("Correlation Heatmap")
    plt.tight_layout()

    img_path = os.path.join(ASSETS_DIR, "correlation_heatmap.png")
    plt.savefig(img_path)
    plt.close()

    return img_path


def insert_imgs(workbook, overview_img_path, correlation_img_path):
    workbook["overview"].add_image(Image(correlation_img_path), "A15")
    workbook["overview"].add_image(Image(overview_img_path), "A45")

def main():
    df = pd.read_csv(DATASET_PATH, delimiter=',', index_col=0)

    print(df.head())

    os.makedirs(ASSETS_DIR, exist_ok=True)

    summary = build_summary(df)
    features = calculate_value_counts(df)

    write_tables(summary, features)
    workbook = load_workbook(OUTPUT_XLSX)
    for feature_name, value_counts in features.items():
        feature_img_path = plot_feature(feature_name, value_counts, df)
        insert_img_into_sheet(workbook, feature_name, feature_img_path)

    overview_img_path = plot_overwiew_pairplot(df)
    correlation_img_path = plot_correlation_heatmap(df)
    insert_imgs(workbook, overview_img_path, correlation_img_path)
    workbook.save(OUTPUT_XLSX)

    # data_audit - https://docs.google.com/spreadsheets/d/1rDzUB413CUG4EwkapDrPBbcwsTbb2mmyU-_1ufRm0gI/edit?usp=sharing


if __name__ == '__main__':
    main()
